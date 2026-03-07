#!/usr/bin/env python3
import argparse
import csv
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

IP_RE = re.compile(r'^(\d{1,3}\.){3}\d{1,3}$')
MAC_RE = re.compile(r'^[0-9A-Fa-f:]{17}$')

def parse_devices_txt(path):
    rows = {}
    with open(path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) >= 3 and IP_RE.match(parts[0]) and MAC_RE.match(parts[1]):
                rows[parts[0]] = {"mac": parts[1].lower(), "vendor": " ".join(parts[2:]).strip()}
    return rows

def parse_nmap_xml(path):
    tree = ET.parse(path)
    root = tree.getroot()
    rows = {}
    for host in root.findall("host"):
        status = host.find("status")
        if status is None or status.get("state") != "up":
            continue
        ipv4 = ""
        mac = ""
        vendor = ""
        for addr in host.findall("address"):
            if addr.get("addrtype") == "ipv4":
                ipv4 = addr.get("addr", "")
            elif addr.get("addrtype") == "mac":
                mac = addr.get("addr", "").lower()
                vendor = addr.get("vendor", "")
        if not ipv4:
            continue
        hostnames = [hn.get("name") for hn in host.findall("./hostnames/hostname") if hn.get("name")]
        ports = []
        services = []
        for p in host.findall("./ports/port"):
            state = p.find("state")
            if state is None or state.get("state") != "open":
                continue
            portid = p.get("portid", "")
            proto = p.get("protocol", "")
            svc = p.find("service")
            name = svc.get("name", "") if svc is not None else ""
            product = svc.get("product", "") if svc is not None else ""
            version = svc.get("version", "") if svc is not None else ""
            extra = " ".join(x for x in [product, version] if x).strip()
            ports.append(f"{portid}/{proto}")
            services.append(f"{portid}:{name}" + (f" ({extra})" if extra else ""))
        os_guess = ""
        osmatch = host.find("./os/osmatch")
        if osmatch is not None:
            os_guess = osmatch.get("name", "")
        rows[ipv4] = {
            "hostnames": ", ".join(hostnames),
            "mac": mac,
            "vendor": vendor,
            "open_ports": ", ".join(ports),
            "services": "; ".join(services),
            "os_guess": os_guess,
        }
    return rows

def map_remote_ip(local_ip, mapped_prefix):
    if not mapped_prefix:
        return ""
    return f"{mapped_prefix}.{local_ip.split('.')[-1]}"

def guess_device(vendor, open_ports, services):
    text = f"{vendor} {open_ports} {services}".lower()
    if "peplink" in text:
        return "Router"
    if "abb" in text or "totalflow" in text:
        return "Flow computer / RTU"
    if "22/tcp" in open_ports and "9090/tcp" in open_ports:
        return "Edge computer"
    return ""

def build_rows(site, devices_map, nmap_map, mapped_prefix):
    ips = sorted(set(devices_map) | set(nmap_map), key=lambda s: tuple(int(x) for x in s.split(".")))
    out = []
    for ip in ips:
        d = devices_map.get(ip, {})
        n = nmap_map.get(ip, {})
        vendor = d.get("vendor") or n.get("vendor") or ""
        mac = d.get("mac") or n.get("mac") or ""
        open_ports = n.get("open_ports", "")
        services = n.get("services", "")
        out.append({
            "site": site,
            "remote_ip": map_remote_ip(ip, mapped_prefix),
            "local_ip": ip,
            "hostname": n.get("hostnames", ""),
            "mac": mac,
            "manufacturer": vendor,
            "open_ports": open_ports,
            "services": services,
            "os_guess": n.get("os_guess", ""),
            "device_guess": guess_device(vendor, open_ports, services),
            "notes": "",
        })
    return out

def write_csv(rows, path):
    fields = ["site","remote_ip","local_ip","hostname","mac","manufacturer","open_ports","services","os_guess","device_guess","notes"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    headers = ["Site","Remote IP","Local IP","Hostname","MAC","Manufacturer","Open Ports","Services","OS Guess","Device Guess","Notes"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r[h.lower().replace(" ","_")] if h not in ("OS Guess","Device Guess") else (r["os_guess"] if h=="OS Guess" else r["device_guess"]) for h in headers])
    from openpyxl.utils import get_column_letter
    widths = [14,14,14,20,20,28,20,42,20,22,20]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("site")
    ap.add_argument("devices_txt")
    ap.add_argument("services_xml")
    ap.add_argument("--mapped-prefix", default="")
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()
    rows = build_rows(args.site, parse_devices_txt(args.devices_txt), parse_nmap_xml(args.services_xml), args.mapped_prefix)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    stem = re.sub(r'[^A-Za-z0-9._-]+', '_', args.site)
    csv_path = outdir / f"{stem}_inventory.csv"
    xlsx_path = outdir / f"{stem}_inventory.xlsx"
    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)
    print(f"Wrote {csv_path}")
    print(f"Wrote {xlsx_path}")

if __name__ == "__main__":
    main()
